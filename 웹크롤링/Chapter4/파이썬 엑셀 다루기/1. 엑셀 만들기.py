import openpyxl

# 엑셀 만들기

wb = openpyxl.Workbook()

# 엑셀 워크시트 만들기

ws = wb.create_sheet('오징어 게임')

# 데이터 추가하기

ws['A1'] = '참가 번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '오일남'

# 엑셀 저장하기

wb.save(r'C:\Users\yongs\GDSC\StudyPython\웹크롤링\Chapter4\파이썬 엑셀 다루기\참가자_data.xlsx') #이스케이프 문자 x just 문자로! r = row